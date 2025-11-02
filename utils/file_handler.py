"""
SME - File handling utilities (Excel and CSV)
"""
import csv
from typing import List, Callable, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from config.constants import EXCEL_MAX_ROWS, CSV_MAX_ROWS


class FileHandler:
    """Handles Excel and CSV file creation with automatic splitting"""
    
    def __init__(self, log_callback: Optional[Callable] = None):
        """
        Initialize file handler
        
        Args:
            log_callback: Optional callback function for logging
        """
        self.log_callback = log_callback
    
    def _log(self, message: str):
        """Internal logging helper"""
        if self.log_callback:
            self.log_callback(message)
    
    def create_excel_file(self, rows: List[List[str]], output_path: str) -> str:
        """
        Create Excel file with automatic sheet creation when hitting row limits
        
        Args:
            rows: List of rows to write (first row should be header)
            output_path: Path to save the Excel file
            
        Returns:
            Path to created file
        """
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        
        current_sheet_num = 1
        current_row = 0
        ws = None
        
        for row_data in rows:
            # Create new sheet if needed
            if current_row == 0 or current_row >= EXCEL_MAX_ROWS - 1:
                if current_sheet_num == 1:
                    ws = wb.create_sheet("Picklist Export")
                else:
                    ws = wb.create_sheet(f"Picklist Export ({current_sheet_num})")
                    self._log(f"  Creating additional sheet: Picklist Export ({current_sheet_num})")
                
                current_sheet_num += 1
                current_row = 0
                
                # Add header to new sheet
                ws.append(rows[0])
                self._apply_header_formatting(ws)
                current_row += 1
                
                # Skip header row in data
                if row_data == rows[0]:
                    continue
            
            ws.append(row_data)
            current_row += 1
        
        # Apply column width to all sheets
        for sheet in wb.worksheets:
            self._apply_column_width(sheet)
            sheet.freeze_panes = "A2"
        
        wb.save(output_path)
        self._log(f"✅ Excel file created: {output_path}")
        self._log(f"✅ Total sheets: {len(wb.worksheets)}")
        self._log(f"✅ Total data rows: {len(rows) - 1}")
        self._log("=" * 70)
        
        return output_path
    
    def _apply_header_formatting(self, ws):
        """Apply formatting to header row"""
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    def _apply_column_width(self, ws):
        """Apply optimal column width"""
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    max_length = max(max_length, len(str(cell.value)) if cell.value else 0)
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
    
    def create_csv_file(self, rows: List[List[str]], output_path: str) -> str:
        """
        Create CSV file(s) - split into multiple files if needed
        
        Args:
            rows: List of rows to write (first row should be header)
            output_path: Path to save the CSV file
            
        Returns:
            Path to first created file
        """
        base_path = output_path.rsplit('.', 1)[0]
        file_num = 1
        current_row = 0
        files_created = []
        
        csv_file = None
        csv_writer = None
        
        for row_data in rows:
            # Create new CSV file every CSV_MAX_ROWS
            if current_row == 0 or current_row >= CSV_MAX_ROWS:
                if csv_file:
                    csv_file.close()
                
                if file_num == 1:
                    file_path = f"{base_path}.csv"
                else:
                    file_path = f"{base_path}_Part{file_num}.csv"
                    self._log(f"  Creating additional CSV file: Part {file_num}")
                
                files_created.append(file_path)
                csv_file = open(file_path, 'w', newline='', encoding='utf-8')
                csv_writer = csv.writer(csv_file)
                
                # Write header to new file
                csv_writer.writerow(rows[0])
                current_row = 1
                file_num += 1
                
                # Skip header row in data
                if row_data == rows[0]:
                    continue
            
            csv_writer.writerow(row_data)
            current_row += 1
        
        if csv_file:
            csv_file.close()
        
        self._log(f"✅ CSV file(s) created: {len(files_created)} file(s)")
        for f in files_created:
            self._log(f"   - {f}")
        self._log(f"✅ Total data rows: {len(rows) - 1}")
        self._log("=" * 70)
        
        return files_created[0] if files_created else output_path