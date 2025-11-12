"""
SME - SOQL Query Runner Module
Executes SOQL queries and exports results
"""
import csv
from typing import List, Dict, Optional, Callable
from core.salesforce_client import SalesforceClient


class SOQLQueryRunner:
    """Handles SOQL query execution and CSV export"""
    
    def __init__(self, sf_client: SalesforceClient, status_callback: Optional[Callable] = None):
        """
        Initialize SOQL query runner
        
        Args:
            sf_client: Salesforce client instance
            status_callback: Optional callback for status updates
        """
        self.sf_client = sf_client
        self.status_callback = status_callback
    
    def _log_status(self, message: str):
        """Internal logging helper"""
        if self.status_callback:
            self.status_callback(message)
    
    def execute_query(self, query: str) -> List[Dict]:
        """
        Execute SOQL query and return results
        
        Args:
            query: SOQL query string
            
        Returns:
            List of dictionaries representing query results
            
        Raises:
            Exception: If query execution fails
        """
        try:
            # Clean the query
            query = query.strip()
            
            # Validate basic SOQL syntax
            if not query.upper().startswith('SELECT'):
                raise ValueError("Query must start with SELECT")
            
            self._log_status(f"Executing query: {query[:100]}...")
            
            # Execute query using Salesforce API
            results = self._execute_soql_query(query)
            
            self._log_status(f"Query returned {len(results)} records")
            
            return results
        
        except Exception as e:
            error_msg = str(e)
            self._log_status(f"Query execution error: {error_msg}")
            raise Exception(f"Query execution failed: {error_msg}")
    
    def _execute_soql_query(self, query: str) -> List[Dict]:
        """
        Execute SOQL query using REST API
        
        Args:
            query: SOQL query string
            
        Returns:
            List of record dictionaries
        """
        all_records = []
        
        try:
            # Use Salesforce REST API query endpoint
            url = f"{self.sf_client.base_url}/services/data/v{self.sf_client.api_version}/query/"
            response = self.sf_client.make_api_call(url, params={'q': query}, method="GET")
            
            if not response or response.status_code != 200:
                error_data = response.json() if response and response.content else {}
                error_message = "Unknown error"
                
                if error_data:
                    if isinstance(error_data, list) and len(error_data) > 0:
                        error_message = error_data[0].get('message', 'Unknown error')
                    elif isinstance(error_data, dict):
                        error_message = error_data.get('message', 'Unknown error')
                
                raise Exception(f"API Error: {error_message}")
            
            data = response.json()
            records = data.get('records', [])
            
            # Flatten records (remove attributes and handle relationships)
            for record in records:
                flattened = self._flatten_record(record)
                all_records.append(flattened)
            
            # Handle pagination if there are more records
            next_records_url = data.get('nextRecordsUrl')
            while next_records_url:
                self._log_status(f"Fetching next batch... (Total so far: {len(all_records)})")
                
                full_url = f"{self.sf_client.base_url}{next_records_url}"
                response = self.sf_client.make_api_call(full_url, method="GET")
                
                if response and response.status_code == 200:
                    data = response.json()
                    records = data.get('records', [])
                    
                    for record in records:
                        flattened = self._flatten_record(record)
                        all_records.append(flattened)
                    
                    next_records_url = data.get('nextRecordsUrl')
                else:
                    break
            
            return all_records
        
        except Exception as e:
            raise Exception(f"Query execution error: {str(e)}")
    
    def _flatten_record(self, record: Dict) -> Dict:
        """
        Flatten a Salesforce record, handling nested relationships
        
        Args:
            record: Salesforce record dictionary
            
        Returns:
            Flattened dictionary
        """
        flattened = {}
        
        for key, value in record.items():
            # Skip the 'attributes' metadata
            if key == 'attributes':
                continue
            
            # Handle nested objects (relationships)
            if isinstance(value, dict):
                if 'attributes' in value:
                    # It's a relationship - flatten it
                    for nested_key, nested_value in value.items():
                        if nested_key != 'attributes':
                            # Create dotted notation: Account.Name
                            flattened[f"{key}.{nested_key}"] = nested_value
                else:
                    # Regular nested dict
                    flattened[key] = str(value)
            elif value is None:
                flattened[key] = ''
            else:
                flattened[key] = value
        
        return flattened
    
    def export_to_csv(self, results: List[Dict], output_path: str):
        """
        Export query results to CSV file
        
        Args:
            results: List of record dictionaries
            output_path: Path to save CSV file
            
        Raises:
            Exception: If export fails
        """
        try:
            if not results:
                raise ValueError("No results to export")
            
            self._log_status(f"Exporting {len(results)} records to CSV...")
            
            # Get all unique column names from all records
            all_columns = set()
            for record in results:
                all_columns.update(record.keys())
            
            # Sort columns alphabetically
            columns = sorted(list(all_columns))
            
            # Write CSV
            with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=columns, extrasaction='ignore')
                
                # Write header
                writer.writeheader()
                
                # Write rows
                for record in results:
                    # Ensure all columns are present (fill missing with empty string)
                    row = {col: record.get(col, '') for col in columns}
                    writer.writerow(row)
            
            self._log_status(f"✅ CSV export complete: {output_path}")
        
        except Exception as e:
            raise Exception(f"CSV export failed: {str(e)}")
        
    def export_to_excel(self, results: List[Dict], output_path: str):
        """
        Export query results to Excel file
        
        Args:
            results: List of record dictionaries
            output_path: Path to save Excel file
            
        Raises:
            Exception: If export fails
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            if not results:
                raise ValueError("No results to export")
            
            self._log_status(f"Exporting {len(results)} records to Excel...")
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Query Results"
            
            # Get all unique column names
            all_columns = set()
            for record in results:
                all_columns.update(record.keys())
            
            columns = sorted(list(all_columns))
            
            # Write header row
            for col_idx, col_name in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Write data rows
            for row_idx, record in enumerate(results, 2):
                for col_idx, col_name in enumerate(columns, 1):
                    value = record.get(col_name, '')
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Auto-adjust column widths
            for col_idx, col_name in enumerate(columns, 1):
                max_length = len(col_name)
                for row_idx in range(2, min(len(results) + 2, 100)):  # Check first 100 rows
                    cell_value = str(ws.cell(row=row_idx, column=col_idx).value or '')
                    max_length = max(max_length, len(cell_value))
                
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 50)
            
            # Freeze header row
            ws.freeze_panes = "A2"
            
            # Save workbook
            wb.save(output_path)
            
            self._log_status(f"✅ Excel export complete: {output_path}")
        
        except Exception as e:
            raise Exception(f"Excel export failed: {str(e)}")